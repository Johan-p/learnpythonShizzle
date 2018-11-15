"""
Excel automation with Python
"""
print(__doc__)

import openpyxl
from openpyxl.styles import Font

wb = openpyxl.load_workbook("example.xlsx")
type(wb)

#getting the sheetnames from the file (workbook tabs)
wb.sheetnames

#wb[sheetname]
sheet = wb['Sheet1']
print type(sheet)

#target the cell
print sheet['A1']

#returns the value in a cell
print sheet['A1'].value

#changing the value in a cell
sheet['A1'] = 'Hello World'
print sheet['A1'].value

#creating worksheets
wb.create_sheet("test")

#removing worksheets
wb.remove_sheet(wb['test'])

#creating worksheet on a specific place (for example the first tab)
wb.create_sheet(index=0, title="First Sheet")

print wb.sheetnames

#returns all the records from c1 to c7
#note that the range starts at 1
for i in range(1, 5): 
	print(sheet.cell(row=i, column=3).value)

#To shows the lenght of data in each row/column
#use this to get the entire range of a column or row.
sheet.max_row
sheet.max_column

#changing the look of a cell
sheet['B1'].font = Font(sz=14, bold=True, italic=True)


#to save the changes to a file.
wb.save('example2.xlsx')

def var():
	pass


a = None

if __name__=='__main__':
	var()